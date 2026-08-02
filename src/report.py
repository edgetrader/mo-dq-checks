"""
The Excel report.

Three sheets, in the order someone actually reads them:

  Summary   what happened, at a glance -- verdict banner, run details,
            totals, and a per-check breakdown showing which checks are
            failing across how many tables
  Results   the full matrix: one row per table, one column per check.
            Check headers are rotated so the grid stays narrow enough to
            scan without scrolling
  Issues    everything that isn't a clean pass, as a flat list, failures
            before warnings -- the sheet you work from

Cell values in the matrix are the plain strings PASS/FAIL/WARN, not symbols,
so the sheet stays filterable and machine-readable; the visual weight comes
from colour. Icons are used only where they can't be mistaken for data (the
verdict banner and the row-status column).

A warning is a finding that shouldn't fail the job -- it is coloured amber
throughout and never affects the exit code.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from checks import CHECK_NAMES, CheckResult

# --- palette ---------------------------------------------------------------
NAVY = "1F4E79"
SLATE = "595959"
RULE = "D9D9D9"

PASS_BG, PASS_FG = "C6EFCE", "006100"
FAIL_BG, FAIL_FG = "FFC7CE", "9C0006"
WARN_BG, WARN_FG = "FFEB9C", "9C6500"
# Mid grey, not near-white: a "didn't run" cell must read as a deliberate
# state rather than looking like an empty cell someone forgot to fill.
SKIP_BG, SKIP_FG = "F2F2F2", "808080"
BAND_BG = "F7F9FC"

TAB_OK, TAB_WARN, TAB_BAD = "00B050", "FFC000", "C00000"

PASS_ICON, WARN_ICON, FAIL_ICON, SKIP_TEXT = "✔", "!", "✖", "n/a"

# --- reusable styles -------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=NAVY)
SUBTITLE_FONT = Font(size=10, italic=True, color=SLATE)
SECTION_FONT = Font(size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor=NAVY)
LABEL_FONT = Font(size=10, bold=True, color=SLATE)
VALUE_FONT = Font(size=10)
HEADER_FONT = Font(size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
# Results data cells are middle-aligned so a row stays visually level even
# when a wrapped comment makes it several lines tall.
LEFT_MIDDLE = Alignment(horizontal="left", vertical="center")
LEFT_MIDDLE_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ROTATED = Alignment(textRotation=90, horizontal="center", vertical="bottom", wrap_text=False)

THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CHECK_COL_WIDTH = 9
COMMENTS_WIDTH = 62
PATH_WIDTH = 58


@dataclass
class ReportData:
    """Everything the sheets need, derived once from the raw results."""

    yyyymm: str
    data_root: str
    run_time: datetime
    tables: list[str] = field(default_factory=list)
    checks: list[str] = field(default_factory=list)
    by_table: dict[str, dict[str, CheckResult]] = field(default_factory=dict)
    file_paths: dict[str, str] = field(default_factory=dict)

    @property
    def all_results(self) -> list[CheckResult]:
        return [r for table in self.tables for r in self.by_table[table].values()]

    @property
    def failures(self) -> list[CheckResult]:
        return [r for r in self.all_results if r.status == "FAIL"]

    @property
    def warnings(self) -> list[CheckResult]:
        return [r for r in self.all_results if r.status == "WARN"]

    @property
    def findings(self) -> list[CheckResult]:
        """Everything that isn't a clean pass, failures first."""
        return self.failures + self.warnings

    @property
    def total(self) -> int:
        return len(self.all_results)

    @property
    def passed(self) -> int:
        return self.total - len(self.findings)

    @property
    def failing_tables(self) -> list[str]:
        return [t for t in self.tables if any(r.status == "FAIL" for r in self.by_table[t].values())]

    @property
    def warning_tables(self) -> list[str]:
        """Tables with warnings but no outright failure."""
        return [
            t for t in self.tables
            if t not in self.failing_tables
            and any(r.status == "WARN" for r in self.by_table[t].values())
        ]

    @property
    def ok(self) -> bool:
        return not self.failures

    def status_of(self, table: str, check: str) -> str | None:
        result = self.by_table[table].get(check)
        return result.status if result else None

    def comments_for(self, table: str) -> str:
        results = self.by_table[table]
        return "\n".join(
            f"{check}: {results[check].message}" for check in self.checks
            if check in results and results[check].message
        )

    def check_stats(self, check: str) -> tuple[int, int, int]:
        """(tables the check ran on, tables it failed, tables it warned on)."""
        ran = sum(1 for t in self.tables if check in self.by_table[t])
        failed = sum(1 for t in self.tables if self.status_of(t, check) == "FAIL")
        warned = sum(1 for t in self.tables if self.status_of(t, check) == "WARN")
        return ran, failed, warned


def organise(yyyymm: str, data_root: str, results: list[CheckResult], run_time: datetime) -> ReportData:
    data = ReportData(yyyymm=yyyymm, data_root=data_root, run_time=run_time)

    for result in results:
        if result.table_name not in data.by_table:
            data.by_table[result.table_name] = {}
            data.tables.append(result.table_name)
        data.by_table[result.table_name][result.check_name] = result

    emitted = {r.check_name for r in results}
    data.checks = [c for c in CHECK_NAMES if c in emitted]
    # Anything emitted that CHECK_NAMES doesn't know about still gets a
    # column rather than vanishing (the test suite treats that as a failure,
    # but a live run must never silently drop a result).
    data.checks += sorted(emitted - set(data.checks))

    # The file path is carried in the file_exists message. read_table now
    # exposes it properly as LoadedTable.path, but run_checks_for_table
    # returns only results, so it's still recovered here.
    for table in data.tables:
        result = data.by_table[table].get("file_exists")
        if result:
            prefix = "File not found: "
            message = result.message
            data.file_paths[table] = (
                message[len(prefix):] if message.startswith(prefix) else message
            )

    return data


def write_excel_report(
    yyyymm: str,
    data_root: str,
    results: list[CheckResult],
    output_dir: str,
    run_time: datetime,
) -> Path:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"dq_check_report_{yyyymm}_{run_time:%Y%m%d_%H%M%S}.xlsx"

    data = organise(yyyymm, str(Path(data_root).resolve()), results, run_time)

    workbook = Workbook()
    _summary_sheet(workbook.active, data)
    _results_sheet(workbook.create_sheet("Results"), data)
    _issues_sheet(workbook.create_sheet("Issues"), data)
    workbook.save(path)

    return path


# --------------------------------------------------------------------------
# Summary
# --------------------------------------------------------------------------


def _tab_colour(data: ReportData) -> str:
    if data.failures:
        return TAB_BAD
    return TAB_WARN if data.warnings else TAB_OK


def _summary_sheet(ws, data: ReportData) -> None:
    ws.title = "Summary"
    ws.sheet_properties.tabColor = _tab_colour(data)
    ws.sheet_view.showGridLines = False

    for column, width in zip("ABCD", (26, 34, 12, 14)):
        ws.column_dimensions[column].width = width

    ws["A1"] = "DATA QUALITY REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = "Monthly Excel table exports"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")
    ws.row_dimensions[1].height = 26

    # Verdict banner -- the one thing to see on opening the file. Warnings
    # get their own state: they don't fail the job, but "all passed" would be
    # a lie when something was flagged.
    if data.failures:
        verdict = (
            f"{FAIL_ICON}   {len(data.failures)} OF {data.total} CHECKS FAILED "
            f"ACROSS {len(data.failing_tables)} TABLE(S)"
        )
        if data.warnings:
            verdict += f", {len(data.warnings)} WARNING(S)"
        background, foreground = FAIL_BG, FAIL_FG
    elif data.warnings:
        verdict = (
            f"{WARN_ICON}   NO FAILURES, BUT {len(data.warnings)} WARNING(S) "
            f"ACROSS {len(data.warning_tables)} TABLE(S)"
        )
        background, foreground = WARN_BG, WARN_FG
    else:
        verdict = f"{PASS_ICON}   ALL {data.total} CHECKS PASSED"
        background, foreground = PASS_BG, PASS_FG

    ws["A4"] = verdict
    ws["A4"].font = Font(size=13, bold=True, color=foreground)
    ws["A4"].fill = PatternFill("solid", fgColor=background)
    ws["A4"].alignment = CENTER
    ws.merge_cells("A4:D4")
    ws.row_dimensions[4].height = 30

    row = _section(ws, 6, "RUN DETAILS")
    for label, value in (
        ("Reporting month", data.yyyymm),
        ("Data root", data.data_root),
        ("Run at", data.run_time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Tables checked", len(data.tables)),
    ):
        row = _label_value(ws, row, label, value)

    row = _section(ws, row + 1, "CHECK TOTALS")
    row = _label_value(ws, row, "Checks run", data.total)
    row = _label_value(ws, row, "Passed", data.passed, colour=PASS_FG)
    row = _label_value(ws, row, "Failed", len(data.failures), colour=FAIL_FG if data.failures else None)
    row = _label_value(ws, row, "Warnings", len(data.warnings), colour=WARN_FG if data.warnings else None)
    row = _label_value(ws, row, "Tables with failures", len(data.failing_tables))

    _check_breakdown(ws, row + 1, data)


def _section(ws, row: int, title: str) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for column in range(2, 5):
        ws.cell(row=row, column=column).fill = SECTION_FILL
    ws.row_dimensions[row].height = 18
    return row + 1


def _label_value(ws, row: int, label: str, value, colour: str | None = None) -> int:
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    cell = ws.cell(row=row, column=2, value=value)
    cell.font = Font(size=10, bold=colour is not None, color=colour) if colour else VALUE_FONT
    return row + 1


def _check_breakdown(ws, start: int, data: ReportData) -> None:
    """Which checks are failing, and how widely -- spots systemic problems."""
    _section(ws, start, "BY CHECK")

    note = ws.cell(
        row=start + 1,
        column=1,
        value=f"Clean rate is over all {len(data.tables)} tables checked.",
    )
    note.font = SUBTITLE_FONT
    ws.merge_cells(start_row=start + 1, start_column=1, end_row=start + 1, end_column=4)

    header_row = start + 2
    for column, title in enumerate(("Check", "Tables failed", "Tables warned", "Clean rate"), start=1):
        cell = ws.cell(row=header_row, column=column, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BOX

    # Deliberately measured against every table checked, not against the
    # tables the check emitted a result for. Several checks don't apply to
    # every table -- mandate_id_completeness, kpi_completeness,
    # source_timeliness -- so an "of those that ran" rate would swing wildly
    # on a small denominator and read as alarming when little is wrong.
    total_tables = len(data.tables) or 1

    row = header_row + 1
    for check in data.checks:
        _ran, failed, warned = data.check_stats(check)
        ws.cell(row=row, column=1, value=check).font = VALUE_FONT

        for column, count, background, foreground in (
            (2, failed, FAIL_BG, FAIL_FG),
            (3, warned, WARN_BG, WARN_FG),
        ):
            cell = ws.cell(row=row, column=column, value=count)
            cell.alignment = CENTER
            if count:
                cell.font = Font(size=10, bold=True, color=foreground)
                cell.fill = PatternFill("solid", fgColor=background)

        # "Clean" means neither failed nor warned.
        rate = ws.cell(row=row, column=4, value=(total_tables - failed - warned) / total_tables)
        rate.number_format = "0%"
        rate.alignment = CENTER
        for column in range(1, 5):
            ws.cell(row=row, column=column).border = BOX
        row += 1

    if row > header_row + 1:
        ws.conditional_formatting.add(
            f"D{header_row + 1}:D{row - 1}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"),
        )


# --------------------------------------------------------------------------
# Results matrix
# --------------------------------------------------------------------------


def _results_sheet(ws, data: ReportData) -> None:
    ws.sheet_properties.tabColor = _tab_colour(data)
    ws.sheet_view.showGridLines = False

    headers = ["", "Table", *data.checks, "Comments", "File"]
    for column, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BOX
        # Rotate the check headers so their columns can stay narrow; the
        # whole matrix then fits on screen without horizontal scrolling.
        cell.alignment = ROTATED if 2 < column <= len(data.checks) + 2 else CENTER
    ws.row_dimensions[1].height = 135

    for offset, table in enumerate(data.tables):
        row = offset + 2
        failed = table in data.failing_tables
        warned = table in data.warning_tables
        banded = PatternFill("solid", fgColor=BAND_BG) if offset % 2 else None

        symbol, colour = (
            (FAIL_ICON, FAIL_FG) if failed
            else (WARN_ICON, WARN_FG) if warned
            else (PASS_ICON, PASS_FG)
        )
        icon = ws.cell(row=row, column=1, value=symbol)
        icon.font = Font(size=12, bold=True, color=colour)
        icon.alignment = CENTER

        name = ws.cell(row=row, column=2, value=table)
        name.font = Font(size=10, bold=failed or warned)
        name.alignment = LEFT_MIDDLE
        if banded:
            name.fill = banded

        for index, check in enumerate(data.checks):
            column = index + 3
            status = data.status_of(table, check)
            cell = ws.cell(row=row, column=column, value=status or SKIP_TEXT)
            cell.alignment = CENTER
            cell.border = BOX
            if status == "PASS":
                cell.fill = PatternFill("solid", fgColor=PASS_BG)
                cell.font = Font(size=9, color=PASS_FG)
            elif status == "FAIL":
                cell.fill = PatternFill("solid", fgColor=FAIL_BG)
                cell.font = Font(size=9, bold=True, color=FAIL_FG)
            elif status == "WARN":
                cell.fill = PatternFill("solid", fgColor=WARN_BG)
                cell.font = Font(size=9, bold=True, color=WARN_FG)
            else:
                # Never ran: not applicable to this table, or an earlier
                # structural failure stopped the chain.
                cell.fill = PatternFill("solid", fgColor=SKIP_BG)
                cell.font = Font(size=9, color=SKIP_FG)

        comments_col = len(data.checks) + 3
        comment = ws.cell(row=row, column=comments_col, value=data.comments_for(table))
        comment.alignment = LEFT_MIDDLE_WRAP
        comment.font = Font(size=9)
        if banded:
            comment.fill = banded

        path = ws.cell(row=row, column=comments_col + 1, value=data.file_paths.get(table, ""))
        path.alignment = LEFT_MIDDLE
        path.font = Font(size=8, color=SLATE)
        if banded:
            path.fill = banded

    last_column = len(headers)
    last_row = len(data.tables) + 1
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    for index in range(len(data.checks)):
        ws.column_dimensions[get_column_letter(index + 3)].width = CHECK_COL_WIDTH
    ws.column_dimensions[get_column_letter(len(data.checks) + 3)].width = COMMENTS_WIDTH
    ws.column_dimensions[get_column_letter(len(data.checks) + 4)].width = PATH_WIDTH

    ws.freeze_panes = "C2"  # keep status + table name visible when scrolling
    ws.auto_filter.ref = f"A1:{get_column_letter(last_column)}{max(last_row, 1)}"

    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"


# --------------------------------------------------------------------------
# Issues
# --------------------------------------------------------------------------


def _issues_sheet(ws, data: ReportData) -> None:
    """
    Everything that isn't a clean pass, as a flat list.

    Failures first, then warnings -- a warning doesn't fail the job but still
    wants looking at, and burying it in the matrix would hide it.
    """
    ws.sheet_properties.tabColor = _tab_colour(data)
    ws.sheet_view.showGridLines = False

    for column, (title, width) in enumerate(
        (("Severity", 11), ("Table", 26), ("Check", 26), ("What went wrong", 104)), start=1
    ):
        cell = ws.cell(row=1, column=column, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BOX
        ws.column_dimensions[get_column_letter(column)].width = width

    findings = [
        (table, result)
        for status in ("FAIL", "WARN")
        for table in data.tables
        for check in data.checks
        for result in [data.by_table[table].get(check)]
        if result and result.status == status
    ]

    if not findings:
        cell = ws.cell(row=2, column=1, value=f"{PASS_ICON}  No issues — every check passed.")
        cell.font = Font(size=11, bold=True, color=PASS_FG)
        cell.fill = PatternFill("solid", fgColor=PASS_BG)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.row_dimensions[2].height = 24
        return

    for row, (table, result) in enumerate(findings, start=2):
        is_failure = result.status == "FAIL"
        background, foreground = (FAIL_BG, FAIL_FG) if is_failure else (WARN_BG, WARN_FG)

        severity = ws.cell(row=row, column=1, value=result.status)
        severity.font = Font(size=10, bold=True, color=foreground)
        severity.fill = PatternFill("solid", fgColor=background)
        severity.alignment = CENTER

        # Table and check are single values, so they sit level with a
        # detail message that has wrapped over several lines.
        name = ws.cell(row=row, column=2, value=table)
        name.font = Font(size=10, bold=True)
        name.alignment = LEFT_MIDDLE

        failed_check = ws.cell(row=row, column=3, value=result.check_name)
        failed_check.font = Font(size=10, color=foreground)
        failed_check.alignment = LEFT_MIDDLE

        detail = ws.cell(row=row, column=4, value=result.message)
        detail.alignment = LEFT_TOP
        detail.font = Font(size=9)

        for column in range(1, 5):
            ws.cell(row=row, column=column).border = BOX

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(findings) + 1}"
