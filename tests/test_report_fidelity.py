"""
Does the workbook faithfully represent the run?

The other tests check that the *checks* are right. These check that nothing
is lost or distorted on the way into Excel: every result reaches its cell,
every finding reaches the Issues sheet with its message intact, and the
Summary's arithmetic matches the results it claims to summarise.

Worth having separately because a reporting bug is invisible to the check
tests -- the run can be completely correct and the file still wrong.
"""
from __future__ import annotations

import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

import generate_test_data  # noqa: E402
from checks import CHECK_NAMES, run_checks_for_table  # noqa: E402
from config_loader import load_table_configs  # noqa: E402
from report import SKIP_TEXT, write_excel_report  # noqa: E402

CONFIG_PATH = PROJECT_ROOT / "config" / "tables_config.json"

# 202603 has three failures, 202604 includes the stale-source warning, and
# 202701 has no seeded defects at all -- so the all-clean path is covered too.
MONTHS = ["202603", "202604", "202701"]


@pytest.fixture(scope="module")
def reports(tmp_path_factory):
    """Generate the months once, check them, and build their workbooks."""
    root = tmp_path_factory.mktemp("fidelity")
    out = tmp_path_factory.mktemp("reports")
    generate_test_data.main(MONTHS, root=root)
    configs = load_table_configs(str(CONFIG_PATH))

    built = {}
    for yyyymm in MONTHS:
        results = [r for cfg in configs for r in run_checks_for_table(cfg, yyyymm, str(root))]
        path = write_excel_report(yyyymm, str(root), results, str(out), datetime.now())
        built[yyyymm] = (results, load_workbook(path))
    return built


def matrix_of(workbook):
    sheet = workbook["Results"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    rows = {sheet.cell(r, 2).value: r for r in range(2, sheet.max_row + 1)}
    return sheet, headers, rows


@pytest.mark.parametrize("yyyymm", MONTHS)
def test_every_result_reaches_its_cell(reports, yyyymm):
    results, workbook = reports[yyyymm]
    sheet, headers, rows = matrix_of(workbook)

    by_table: dict[str, dict[str, str]] = {}
    for r in results:
        by_table.setdefault(r.table_name, {})[r.check_name] = r.status

    assert set(rows) == set(by_table), "the matrix has the wrong set of table rows"

    for table, statuses in by_table.items():
        for check in CHECK_NAMES:
            if check not in headers:
                continue
            cell = sheet.cell(rows[table], headers.index(check) + 1).value
            # A check that never ran shows the skip marker, never a status.
            assert cell == statuses.get(check, SKIP_TEXT), f"{table}.{check}"


@pytest.mark.parametrize("yyyymm", MONTHS)
def test_issues_sheet_lists_findings_exactly(reports, yyyymm):
    """Nothing missing, and nothing listed that wasn't a finding."""
    results, workbook = reports[yyyymm]
    sheet = workbook["Issues"]

    listed = {
        (sheet.cell(r, 2).value, sheet.cell(r, 3).value, sheet.cell(r, 1).value)
        for r in range(2, sheet.max_row + 1)
        if sheet.cell(r, 1).value in ("FAIL", "WARN")
    }
    expected = {(r.table_name, r.check_name, r.status) for r in results if r.status != "PASS"}

    assert listed == expected


@pytest.mark.parametrize("yyyymm", MONTHS)
def test_finding_messages_survive_verbatim(reports, yyyymm):
    """A truncated or reworded message would hide what actually went wrong."""
    results, workbook = reports[yyyymm]
    sheet = workbook["Issues"]

    shown = {
        (sheet.cell(r, 2).value, sheet.cell(r, 3).value): sheet.cell(r, 4).value
        for r in range(2, sheet.max_row + 1)
        if sheet.cell(r, 1).value in ("FAIL", "WARN")
    }

    for result in results:
        if result.status == "PASS":
            continue
        assert result.message, f"{result.table_name}.{result.check_name} has no message"
        assert shown[(result.table_name, result.check_name)] == result.message


@pytest.mark.parametrize("yyyymm", MONTHS)
def test_comments_carry_every_message_for_the_table(reports, yyyymm):
    results, workbook = reports[yyyymm]
    sheet, headers, rows = matrix_of(workbook)
    comments_col = headers.index("Comments") + 1

    for result in results:
        if not result.message:
            continue
        cell = sheet.cell(rows[result.table_name], comments_col).value or ""
        assert result.message in cell, f"{result.table_name}.{result.check_name} missing"


@pytest.mark.parametrize("yyyymm", MONTHS)
def test_summary_totals_match_the_results(reports, yyyymm):
    results, workbook = reports[yyyymm]
    sheet = workbook["Summary"]
    values = {sheet.cell(r, 1).value: sheet.cell(r, 2).value for r in range(1, 25)}
    counts = Counter(r.status for r in results)

    assert values["Reporting month"] == yyyymm
    assert values["Checks run"] == len(results)
    assert values["Passed"] == counts["PASS"]
    assert values["Failed"] == counts["FAIL"]
    assert values["Warnings"] == counts["WARN"]
    assert values["Tables checked"] == len({r.table_name for r in results})
    assert values["Tables with failures"] == len(
        {r.table_name for r in results if r.status == "FAIL"}
    )


def test_a_clean_month_says_so(reports):
    """202701 has no seeded defects -- the report should be unambiguous."""
    results, workbook = reports["202701"]

    assert not [r for r in results if r.status != "PASS"]
    assert "ALL" in workbook["Summary"]["A4"].value
    assert "No issues" in workbook["Issues"]["A2"].value
