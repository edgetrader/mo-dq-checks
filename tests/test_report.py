"""
Tests for the Excel report.

Formatting is mostly a matter of taste, so these check the things that
would actually mislead someone reading the file: the verdict, the matrix
contents, which cells are marked failed, and the numbers on the summary.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from checks import CHECK_NAMES, CheckResult  # noqa: E402
from report import (  # noqa: E402
    FAIL_BG, FAIL_ICON, PASS_BG, PASS_ICON, WARN_BG, WARN_ICON,
    organise, write_excel_report,
)

RUN_TIME = datetime(2026, 8, 2, 9, 30, 0)


def result(table, check, status, message=""):
    return CheckResult(table, check, status, message)


@pytest.fixture
def clean_results():
    return [
        result("ALPHA", "file_exists", "PASS", "/data/202607/alpha.xlsx"),
        result("ALPHA", "sheet_exists", "PASS"),
        result("ALPHA", "required_columns", "PASS"),
        result("ALPHA", "row_count", "PASS", "4 rows"),
        result("ALPHA", "report_date_dtype", "PASS"),
        result("ALPHA", "val_amt_dtype", "PASS"),
    ]


@pytest.fixture
def mixed_results(clean_results):
    return clean_results + [
        result("BETA", "file_exists", "FAIL", "File not found: /data/202607/beta.xlsx"),
        result("GAMMA", "file_exists", "PASS", "/data/202607/gamma.xlsx"),
        result("GAMMA", "sheet_exists", "PASS"),
        result("GAMMA", "required_columns", "PASS"),
        result("GAMMA", "row_count", "PASS", "9 rows"),
        result("GAMMA", "report_date_dtype", "PASS"),
        result("GAMMA", "val_amt_dtype", "FAIL", "2 non-numeric val_amt value(s)"),
    ]


def write(results, tmp_path, yyyymm="202607"):
    path = write_excel_report(yyyymm, str(tmp_path), results, str(tmp_path / "out"), RUN_TIME)
    return load_workbook(path), path


# --------------------------------------------------------------------------
# organise()
# --------------------------------------------------------------------------


def test_organise_groups_by_table_preserving_run_order(mixed_results):
    data = organise("202607", "/data", mixed_results, RUN_TIME)

    assert data.tables == ["ALPHA", "BETA", "GAMMA"]
    assert data.failing_tables == ["BETA", "GAMMA"]
    assert not data.ok


def test_organise_orders_checks_by_check_names(mixed_results):
    data = organise("202607", "/data", mixed_results, RUN_TIME)
    assert data.checks == [c for c in CHECK_NAMES if c in data.checks]


def test_organise_recovers_the_file_path_either_way(mixed_results):
    """The path is carried in the file_exists message, worded differently on failure."""
    data = organise("202607", "/data", mixed_results, RUN_TIME)

    assert data.file_paths["ALPHA"] == "/data/202607/alpha.xlsx"
    assert data.file_paths["BETA"] == "/data/202607/beta.xlsx"  # prefix stripped


def test_comments_collect_every_message_for_a_table(mixed_results):
    data = organise("202607", "/data", mixed_results, RUN_TIME)
    assert "val_amt_dtype: 2 non-numeric val_amt value(s)" in data.comments_for("GAMMA")


# --------------------------------------------------------------------------
# Workbook structure
# --------------------------------------------------------------------------


def test_report_has_three_sheets(mixed_results, tmp_path):
    workbook, path = write(mixed_results, tmp_path)

    assert workbook.sheetnames == ["Summary", "Results", "Issues"]
    assert path.name == "dq_check_report_202607_20260802_093000.xlsx"


def test_results_matrix_contents(mixed_results, tmp_path):
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Results"]

    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    assert headers[1] == "Table"
    assert headers[-2:] == ["Comments", "File"]

    rows = {sheet.cell(r, 2).value: r for r in range(2, sheet.max_row + 1)}
    assert set(rows) == {"ALPHA", "BETA", "GAMMA"}

    # Row status icons reflect whether the table had any failure.
    assert sheet.cell(rows["ALPHA"], 1).value == PASS_ICON
    assert sheet.cell(rows["BETA"], 1).value == FAIL_ICON

    # Cells keep the plain PASS/FAIL text so the sheet stays filterable.
    val_amt_col = headers.index("val_amt_dtype") + 1
    assert sheet.cell(rows["GAMMA"], val_amt_col).value == "FAIL"
    assert sheet.cell(rows["ALPHA"], val_amt_col).value == "PASS"


def test_failed_cells_are_coloured(mixed_results, tmp_path):
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Results"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    rows = {sheet.cell(r, 2).value: r for r in range(2, sheet.max_row + 1)}
    column = headers.index("val_amt_dtype") + 1

    assert FAIL_BG in sheet.cell(rows["GAMMA"], column).fill.fgColor.rgb
    assert PASS_BG in sheet.cell(rows["ALPHA"], column).fill.fgColor.rgb


def test_checks_that_never_ran_are_blank_not_failed(mixed_results, tmp_path):
    """BETA's file was missing, so nothing downstream ran for it."""
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Results"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    rows = {sheet.cell(r, 2).value: r for r in range(2, sheet.max_row + 1)}

    cell = sheet.cell(rows["BETA"], headers.index("row_count") + 1)
    assert cell.value not in ("PASS", "FAIL")


def test_results_values_are_middle_aligned(mixed_results, tmp_path):
    """Rows stay visually level even when a wrapped comment makes one tall."""
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Results"]

    for row in range(2, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            assert cell.alignment.vertical == "center", f"{cell.coordinate} is not middle-aligned"


def test_results_sheet_is_navigable(mixed_results, tmp_path):
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Results"]

    assert sheet.freeze_panes == "C2"       # status + table stay visible
    assert sheet.auto_filter.ref is not None
    assert sheet.cell(1, 3).alignment.textRotation == 90


# --------------------------------------------------------------------------
# Summary and failures
# --------------------------------------------------------------------------


def test_summary_verdict_and_totals_when_failing(mixed_results, tmp_path):
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Summary"]

    assert sheet["A4"].value.startswith(FAIL_ICON)
    assert "2 OF 13 CHECKS FAILED ACROSS 2 TABLE(S)" in sheet["A4"].value

    values = {sheet.cell(r, 1).value: sheet.cell(r, 2).value for r in range(1, 20)}
    assert values["Reporting month"] == "202607"
    assert values["Tables checked"] == 3
    assert values["Failed"] == 2
    assert values["Tables with failures"] == 2


def test_summary_verdict_when_clean(clean_results, tmp_path):
    workbook, _ = write(clean_results, tmp_path)

    assert workbook["Summary"]["A4"].value.startswith(PASS_ICON)
    assert "ALL 6 CHECKS PASSED" in workbook["Summary"]["A4"].value
    assert workbook["Issues"]["A2"].value.startswith(PASS_ICON)


def test_clean_rate_is_measured_against_every_table(mixed_results, tmp_path):
    """
    analytics_completeness only emits when it fails, so a rate computed over
    "tables that ran it" would read 0% clean whenever it appears at all.
    """
    results = mixed_results + [
        result("ALPHA", "analytics_completeness", "FAIL", "1 null analytics value(s)")
    ]
    workbook, _ = write(results, tmp_path)
    sheet = workbook["Summary"]

    rates = {
        sheet.cell(r, 1).value: sheet.cell(r, 4).value
        for r in range(1, sheet.max_row + 1)
        if isinstance(sheet.cell(r, 4).value, float)
    }
    # One of three tables failed it -> 67% clean, not 0%.
    assert rates["analytics_completeness"] == pytest.approx(2 / 3)


def test_issues_sheet_lists_every_finding(mixed_results, tmp_path):
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Issues"]

    listed = {
        (sheet.cell(r, 1).value, sheet.cell(r, 2).value, sheet.cell(r, 3).value)
        for r in range(2, sheet.max_row + 1)
    }
    assert listed == {
        ("FAIL", "BETA", "file_exists"),
        ("FAIL", "GAMMA", "val_amt_dtype"),
    }
    assert "non-numeric" in sheet.cell(3, 4).value


def test_issues_table_and_check_are_middle_aligned(mixed_results, tmp_path):
    """They should sit level with a detail message that wrapped over several lines."""
    workbook, _ = write(mixed_results, tmp_path)
    sheet = workbook["Issues"]

    for row in range(2, sheet.max_row + 1):
        for column in (2, 3):
            cell = sheet.cell(row, column)
            assert cell.alignment.vertical == "center", f"{cell.coordinate} is not middle-aligned"


def test_tab_colour_signals_the_outcome(clean_results, mixed_results, tmp_path):
    clean, _ = write(clean_results, tmp_path / "a")
    failing, _ = write(mixed_results, tmp_path / "b")

    assert clean["Summary"].sheet_properties.tabColor.rgb != failing["Summary"].sheet_properties.tabColor.rgb


# --------------------------------------------------------------------------
# Warnings
# --------------------------------------------------------------------------


@pytest.fixture
def warned_results(clean_results):
    return clean_results + [
        result("DELTA", "file_exists", "PASS", "/data/202607/delta.xlsx"),
        result("DELTA", "sheet_exists", "PASS"),
        result("DELTA", "source_timeliness", "WARN", "2 mandate(s) sourced from 202606"),
    ]


def test_warning_gets_its_own_verdict(warned_results, tmp_path):
    """A warning isn't a failure, but "all passed" would be a lie."""
    workbook, _ = write(warned_results, tmp_path)
    banner = workbook["Summary"]["A4"].value

    assert banner.startswith(WARN_ICON)
    assert "NO FAILURES, BUT 1 WARNING(S)" in banner


def test_warning_is_counted_separately(warned_results, tmp_path):
    workbook, _ = write(warned_results, tmp_path)
    sheet = workbook["Summary"]
    values = {sheet.cell(r, 1).value: sheet.cell(r, 2).value for r in range(1, 22)}

    assert values["Failed"] == 0
    assert values["Warnings"] == 1
    assert values["Passed"] == 8          # warnings don't count as passes


def test_warned_cell_is_amber_and_flagged(warned_results, tmp_path):
    workbook, _ = write(warned_results, tmp_path)
    sheet = workbook["Results"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    rows = {sheet.cell(r, 2).value: r for r in range(2, sheet.max_row + 1)}
    cell = sheet.cell(rows["DELTA"], headers.index("source_timeliness") + 1)

    assert cell.value == "WARN"
    assert WARN_BG in cell.fill.fgColor.rgb
    assert sheet.cell(rows["DELTA"], 1).value == WARN_ICON


def test_issues_sheet_separates_severities(mixed_results, warned_results, tmp_path):
    workbook, _ = write(mixed_results + warned_results[len(mixed_results) - 7:], tmp_path)
    sheet = workbook["Issues"]
    severities = [sheet.cell(r, 1).value for r in range(2, sheet.max_row + 1)]

    # Failures come before warnings.
    assert severities == sorted(severities, key=lambda s: 0 if s == "FAIL" else 1)
    assert "WARN" in severities
